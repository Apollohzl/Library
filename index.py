import tkinter as tk
from tkinter import messagebox
import json
from datetime import datetime
from tkinter import filedialog
import os
import openpyxl


class Book:
    def __init__(self, title, author, year, borrowed=False):
        self.title = title
        self.author = author
        self.year = year
        self.borrowed = borrowed

class Borrow:
    def __init__(self, title, author, year, borrowtime):
        self.title = title
        self.author = author
        self.year = year
        self.borrowtime = borrowtime

class Library:
    def __init__(self):
        self.books = []
        self.History = []

        self.load_History()
        self.load_books()
    
    def add_book(self, book):
        self.books.append(book)
        #print(f"add book:{book}")
        self.save_books()
    
    def delete_book(self, title):
        self.books = [b for b in self.books if b.title != title]
        self.save_books()
    
    def find_book(self, query):
        results = [b for b in self.books if query.lower() in b.title.lower() or query.lower() in b.author.lower()]
        return results
    
    def list_books(self):
        return self.books
    
    def return_HistoryOfBorrow(self):
        return self.History
    
    def borrow_book(self, title, author, year):
        for book in self.books:
            if book.title == title:
                if book.borrowed:
                    return "这本书已经被借出"
                else:
                    book.borrowed = True
                    self.save_books()
                    self.savebooks_history(title, author, year)
                    return "借书成功"
        return "未找到这本书"
    
    def return_book(self, title):
        for book in self.books:
            if book.title == title:
                if not book.borrowed:
                    return "这本书未被借出"
                else:
                    book.borrowed = False
                    self.save_books()
                    return "还书成功"
        return "未找到这本书"
    
    def save_books(self):
        with open(r'.\books.json', 'w') as file:
            json.dump([book.__dict__ for book in self.books], file)
    
    def savebooks_history(self,title,author="未知",year="未知"):
        now = str(datetime.now())
        borrow_History_data = {"书名：":title, "作者：":author, "出版日期： ":year, "借书日期：":now}
        #print(borrow_History_data)
        self.History.append(borrow_History_data)
        #print(self.History)
        with open(r'.\borrow.json', 'w') as file:
            json.dump([History for History in self.History], file)


    def load_books(self):
        try:
            with open(r'.\books.json', 'r') as file:
                books_data = json.load(file)
                self.books = [Book(**data) for data in books_data]
        except :
            self.books = []
    
    def load_History(self):
        try:
            with open(r'.\borrow.json', 'r') as file:
                borrow_data = json.load(file)
                
                key_mapping = {
                    "\u4e66\u540d\uff1a": 'title',
                    "\u4f5c\u8005\uff1a": 'author',
                    "\u51fa\u7248\u65e5\u671f\uff1a ": 'year',
                    "\u501f\u4e66\u65e5\u671f\uff1a": 'borrowtime'
                }
                
                
                transformed_data = [
                    {key_mapping.get(k, k): v for k, v in data.items()}
                    for data in borrow_data
                ]
                self.History = [Borrow(**data) for data in transformed_data]
                print(self.History)
        except:
            self.History = []
           

class LibraryApp:
    def __init__(self, root):
        self.library = Library()
        self.root = root
        self.root.title("图书馆管理系统")

        self.root.minsize(width=400, height=500)
        self.root.maxsize(width=600, height=700) 

     
        self.title_label = tk.Label(root, text="书名:")
        self.title_entry = tk.Entry(root)
        self.author_label = tk.Label(root, text="作者:")
        self.author_entry = tk.Entry(root)
        self.year_label = tk.Label(root, text="出版年份:")
        self.year_entry = tk.Entry(root)

        self.add_button = tk.Button(root, text="添加书籍", command=self.add_book, bg='lightgreen', fg='black')
        self.delete_button = tk.Button(root, text="删除书籍", command=self.delete_book, bg='#00fbff', fg='black')
        self.find_button = tk.Button(root, text="查找书籍", command=self.find_book, bg='lightblue', fg='black')
        self.list_button = tk.Button(root, text="列出所有书籍", command=self.list_books, bg='#59ff00', fg='black')
        self.borrow_button = tk.Button(root, text="借书", command=self.borrow_book, bg='#7394ff', fg='black')
        self.return_button = tk.Button(root, text="还书", command=self.return_book, bg='#00ffc4', fg='black')
        self.return_HistoryForBorrow_button = tk.Button(root,text="借书历史", command= self.return_HistoryOfBorrowBooks, bg='lightpink', fg='black')
        self.upTheFileOfBook_excel = tk.Button(root,text="一键导入书籍", command=self.open_file, bg='#fbff75', fg='black')
        self.select_folder_button = tk.Button(root, text="一键导出书籍文件", command=self.select_folder, bg='#00FFFF', fg='black')
        self.clean_All_Book = tk.Button(root, text="一键清空所有书籍", command=self.clean_LibraryBook, bg='#FF00FF', fg='black')

        self.results_text = tk.Text(root, height=10, width=50)

        
        self.title_label.grid(row=0, column=0, padx=10, pady=5, sticky='nsew')
        self.title_entry.grid(row=0, column=1, padx=10, pady=5, sticky='nsew')
        self.author_label.grid(row=1, column=0, padx=10, pady=5, sticky='nsew')
        self.author_entry.grid(row=1, column=1, padx=10, pady=5, sticky='nsew')
        self.year_label.grid(row=2, column=0, padx=10, pady=5, sticky='nsew')
        self.year_entry.grid(row=2, column=1, padx=10, pady=5, sticky='nsew')

        self.add_button.grid(row=3, column=0, padx=10, pady=5, sticky='nsew')
        self.delete_button.grid(row=3, column=1, padx=10, pady=5, sticky='nsew')
        self.find_button.grid(row=4, column=0, padx=10, pady=5, sticky='nsew')
        self.list_button.grid(row=4, column=1, padx=10, pady=5, sticky='nsew')
        self.borrow_button.grid(row=5, column=0, padx=10, pady=5, sticky='nsew')
        self.return_button.grid(row=5, column=1, padx=10, pady=5, sticky='nsew')
        self.return_HistoryForBorrow_button.grid(row=6, column=0, padx=20, pady=8, sticky='nsew')
        self.upTheFileOfBook_excel.grid(row=6, column=1, padx=20, pady=8, sticky='nsew')
        self.select_folder_button.grid(row=7, column=0, padx=20, pady=8, sticky='nsew')
        self.clean_All_Book.grid(row=7, column=1, padx=20, pady=8, sticky='nsew')

        self.results_text.grid(row=8, column=0, columnspan=2, padx=10, pady=10, sticky='nsew')

        
        for i in range(8):
            self.root.grid_rowconfigure(i, weight=1)
        for i in range(2):
            self.root.grid_columnconfigure(i, weight=1)

    def clean_LibraryBook(self):
        with open(r'.\books.json','w') as f:
            Empty = []
            json.dump(Empty,f)
        self.clear_result()
        #print("清空书籍")
        messagebox.showinfo("信息", "书籍已全部清空")

    def select_folder(self):
        try:
            folder_selected = filedialog.askdirectory()
            if folder_selected:
                selected_folder_path = folder_selected
                # print(f"选择的文件夹路径: {selected_folder_path}")
            self.out_LibraryBook(selected_folder_path)
            self.list_books()
        except:
            print("未选择文件夹")

    def out_LibraryBook(self,fp):
        #建立工作簿  设置工作表
        file_path = fp+'/outLibraryBook.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'LibraryBook'
        #读取json文件并导出工作簿
        with open('books.json','r') as f:
            try:
                with open(r'.\books.json', 'r') as file:
                    books_data = json.load(file)
                    for book_xinxi in books_data:
                        A = str(book_xinxi['title'])
                        B = str(book_xinxi['author'])
                        C = book_xinxi['year']
                        if A and B and C:
                            # print(f"书名: {book_xinxi['title']}, 作者: {book_xinxi['author']}, 出版年份: {book_xinxi['year']}")
                            row = [A,B,C]
                            ws.append(row)

            except:
                row = []
                ws.append(row)

        wb.save(file_path)
        messagebox.showinfo("信息", "书籍已导出成功")


    def open_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")]
        )
        if file_path:
            #print(f"选择的文件: {file_path}")
            messagebox.showinfo("信息", "书籍已添加成功")

            self.fastAdd_book(file_path)
        else:
            print("没有选择文件")
        

    def add_book(self):
        title = self.title_entry.get()
        author = self.author_entry.get()
        year = self.year_entry.get()
        if title and author and year:
            book = Book(title, author, int(year))
            self.library.add_book(book)
            self.clear_entries()
            messagebox.showinfo("信息", "书籍已添加成功")
        else:
            messagebox.showwarning("警告", "请填写所有字段")
        self.list_books()

    def fastAdd_book(self,fn):
        wb = openpyxl.load_workbook(fn)
        ws = wb['LibraryBook']
        for i in range(1,ws.max_row+1):
            A = str(ws.cell(column=1,row=i).value)
            B = str(ws.cell(column=2,row=i).value)
            C = str(ws.cell(column=3,row=i).value)
            if A!='None' and B!='None' and C!='None':
                book = Book(A, B, int(C))
                self.library.add_book(book)
        self.list_books()

    def delete_book(self):
        title = self.title_entry.get()
        if title:
            self.library.delete_book(title)
            self.clear_entries()
            messagebox.showinfo("信息", "书籍已删除成功")
        else:
            messagebox.showwarning("警告", "请输入书名")
        self.list_books()

    def find_book(self):
        query = self.title_entry.get()
        results = self.library.find_book(query)
        self.display_results(results)

    def list_books(self):
        print("列出书籍")
        books = self.library.list_books()
        print(books)
        self.display_results(books)

    def borrow_book(self):
        title = self.title_entry.get()
        author = self.author_entry.get()
        year = self.year_entry.get()
        if title:
            message = self.library.borrow_book(title,author,year)
            self.clear_entries()
            messagebox.showinfo("信息", message)
        else:
            messagebox.showwarning("警告", "请输入书名")
        self.list_books()

    def return_book(self):
        title = self.title_entry.get()
        print(f"查找书籍 {title}")
        if title:
            message = self.library.return_book(title)
            self.clear_entries()
            messagebox.showinfo("信息", message)
        else:
            messagebox.showwarning("警告", "请输入书名")
        self.list_books()

    def display_results(self, books):
        self.results_text.delete(1.0, tk.END)
        for book in books:
            borrowed_status = "已借出" if book.borrowed else "未借出"
            self.results_text.insert(tk.END, f"书名: {book.title}, 作者: {book.author}, 出版年份: {book.year}, 状态: {borrowed_status}\n")

    def display_History(self,history):
        try:
            self.results_text.delete(1.0, tk.END)
            for History in history:
                #print(f"输出借书记录:{History}")
                self.results_text.insert(tk.END,f"书名: {History['书名：']}, 作者: {History['作者：']}, 出版年份: {History['出版日期： ']}, 借书时间：{History['借书日期：']}")
        except AttributeError:
            raise
            

    def clear_entries(self):
        self.title_entry.delete(0, tk.END)
        self.author_entry.delete(0, tk.END)
        self.year_entry.delete(0, tk.END)
    
    def clear_result(self):
        self.results_text.delete(1.0,tk.END)
    
    def return_HistoryOfBorrowBooks(self):
        BorrowHistoryList = self.library.return_HistoryOfBorrow()
        self.display_History(BorrowHistoryList)

if __name__ == "__main__":
    root = tk.Tk()
    app = LibraryApp(root)
    root.mainloop()
